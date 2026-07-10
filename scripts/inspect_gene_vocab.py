import openpyxl
from pathlib import Path

# Check potential locations of the targets workbook
excel_path = Path("C:/Users/ayu23/OneDrive/Desktop/dock/docking_pipeline/docs/AYUSH_AMR_Final_Targets.xlsx")
if not excel_path.exists():
    excel_path = Path("C:/Users/ayu23/OneDrive/Desktop/dock/docs/AYUSH_AMR_Final_Targets.xlsx")

print("Target workbook found:", excel_path.exists())
if excel_path.exists():
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    print("Worksheet Names:", wb.sheetnames)
    
    sheet = wb.active
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
    print("\nHeaders:", headers)
    
    # Search for MrkH and Wzc rows
    for r_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
        row_str = " ".join([str(x) for x in row if x is not None]).lower()
        if "mrkh" in row_str or "wzc" in row_str:
            print(f"\n[Row {r_idx}] Match Found:")
            for header, val in zip(headers, row):
                print(f"   {header}: {val}")
else:
    print("Authoritative Excel targets workbook is missing!")
