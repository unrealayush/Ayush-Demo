import os
import re
import csv
import json
from typing import List, Dict, Any

# Constants for Master Paths and Legacy CSV Locations
BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
DOCS_DIR = os.path.join(BASE_DIR, "docs")
TARGETS_MASTER = os.path.join(DOCS_DIR, "AYUSH_AMR_Final_Targets.xlsx")
LIGANDS_MASTER = os.path.join(DOCS_DIR, "Verified_AYUSH_Ligands_24 (1).xlsx")

LEGACY_TARGETS_CSV = os.path.join(BASE_DIR, "data", "inputs", "pathogen_target_registry.csv")
LEGACY_LIGANDS_CSV = os.path.join(BASE_DIR, "data", "inputs", "ligand_library.csv")

class RegistryLoader:
    _instance = None

    def __new__(cls):
        if cls._instance is None:
            cls._instance = super(RegistryLoader, cls).__new__(cls)
            cls._instance._init_loader()
        return cls._instance

    def _init_loader(self):
        self.ligands: Dict[str, Dict[str, Any]] = {}
        self.targets: Dict[str, Dict[str, Any]] = {}
        self.load_master_registries()

    def normalize_id(self, name: str) -> str:
        # Strip trailing/leading space, convert to lowercase, replace special chars with underscore
        clean = name.strip().lower()
        clean = re.sub(r"[^a-z0-9]+", "_", clean)
        return clean.strip("_")

    def query_pubchem_smiles(self, name: str) -> str:
        import urllib.request
        import urllib.parse
        import json
        print(f"[PubChem] Fetching SMILES for compound: '{name}'...")
        try:
            encoded_name = urllib.parse.quote(name)
            url = f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/name/{encoded_name}/property/CanonicalSMILES,ConnectivitySMILES/JSON"
            req = urllib.request.Request(url, headers={'User-Agent': 'BioinformaticsHub/1.0'})
            with urllib.request.urlopen(req, timeout=5) as response:
                data = json.loads(response.read().decode('utf-8'))
                properties = data.get("PropertyTable", {}).get("Properties", [])
                if properties:
                    return properties[0].get("CanonicalSMILES") or properties[0].get("ConnectivitySMILES", "None")
        except Exception as e:
            print(f"[PubChem] Error fetching SMILES for '{name}': {e}")
        return "None"

    def load_master_registries(self):
        self.ligands = {}
        self.targets = {}

        # Check Cache first for instant startup (<1s)
        cache_path = os.path.join(BASE_DIR, "data", "inputs", "excel_registries_cache.json")
        if os.path.exists(cache_path):
            try:
                with open(cache_path, "r", encoding="utf-8") as f:
                    cache_data = json.load(f)
                self.ligands = cache_data.get("ligands", {})
                self.targets = cache_data.get("targets", {})
                if self.ligands and self.targets:
                    return
            except Exception as e:
                print(f"[RegistryLoader] Error reading JSON cache: {e}")

        # Lazy import openpyxl ONLY if cache is missing to avoid 1.27s import blocking on startup
        import openpyxl

        # 1. Parse Master Ligands from Excel
        if os.path.exists(LIGANDS_MASTER):
            wb = openpyxl.load_workbook(LIGANDS_MASTER)
            sheet = wb.active # 'Verified_AYUSH_Ligands'
            
            # Read header row
            headers = [cell.value for cell in sheet[1]]
            col_map = {h: idx for idx, h in enumerate(headers) if h is not None}
            
            # Parse rows
            for r in range(2, sheet.max_row + 1):
                row_vals = [cell.value for cell in sheet[r]]
                if not any(row_vals):
                    continue # Skip blank rows
                
                compound = row_vals[col_map["Compound"]]
                if not compound:
                    continue
                
                comp_id = self.normalize_id(compound)
                plant = row_vals[col_map["Plant"]] if "Plant" in col_map else "Unknown Plant"
                chem_class = row_vals[col_map["Chemical Class"]] if "Chemical Class" in col_map else "Unknown Class"
                pubchem = str(row_vals[col_map["PubChem CID"]]) if "PubChem CID" in col_map else "0"
                priority = row_vals[col_map["Priority"]] if "Priority" in col_map else "Low"
                
                # Fetch Canonical SMILES dynamically from PubChem PUG REST API
                smiles = self.query_pubchem_smiles(compound)

                self.ligands[comp_id] = {
                    "compound_id": comp_id,
                    "compound_name": compound.strip(),
                    "plant_name": plant.strip() if plant else "",
                    "chemical_class": chem_class.strip() if chem_class else "",
                    "pubchem_cid": pubchem.strip() if pubchem else "",
                    "priority": priority.strip() if priority else "",
                    "canonical_smiles": smiles
                }
        
        # 2. Parse Master Targets from Excel
        if os.path.exists(TARGETS_MASTER):
            wb = openpyxl.load_workbook(TARGETS_MASTER)
            sheet = wb.active # 'AYUSH_AMR_Final_Targets'
            
            headers = [cell.value for cell in sheet[1]]
            col_map = {h: idx for idx, h in enumerate(headers) if h is not None}
            
            for r in range(2, sheet.max_row + 1):
                row_vals = [cell.value for cell in sheet[r]]
                if not any(row_vals):
                    continue
                
                gene = row_vals[col_map["Gene"]]
                protein = row_vals[col_map["Target Protein"]]
                if not gene or not protein:
                    continue
                
                tgt_id = self.normalize_id(gene)
                organism = row_vals[col_map["Organism"]] if "Organism" in col_map else "Unknown Organism"
                tgt_class = row_vals[col_map["Target Class"]] if "Target Class" in col_map else "Unknown Class"
                function = row_vals[col_map["Primary Function"]] if "Primary Function" in col_map else "Unknown"
                role = row_vals[col_map["Role in AMR/Biofilm"]] if "Role in AMR/Biofilm" in col_map else "Unknown"
                uniprot = row_vals[col_map["UniProt"]] if "UniProt" in col_map else "Unknown"
                avail = str(row_vals[col_map["Structure Availability"]]) if "Structure Availability" in col_map else "None"

                # Extract first PDB code from availability (e.g. "PDB: 4G4K" -> "4G4K")
                pdb_id = "structure_pending"
                pdb_match = re.search(r"PDB:\s*([A-Za-z0-9]+)", avail, re.I)
                if pdb_match:
                    pdb_id = pdb_match.group(1).upper()

                self.targets[tgt_id] = {
                    "target_id": tgt_id,
                    "target_name": protein.strip(),
                    "gene": gene.strip(),
                    "organism": organism.strip(),
                    "target_class": tgt_class.strip() if tgt_class else "",
                    "primary_function": function.strip() if function else "",
                    "role_in_amr": role.strip() if role else "",
                    "uniprot": uniprot.strip() if uniprot else "",
                    "structure_availability": avail.strip(),
                    "pdb_id": pdb_id
                }

        # 3. Export Legacy CSV Backwards Compatibility automatically
        self.generate_legacy_csvs()

        # Save Cache for subsequent instant startups
        try:
            os.makedirs(os.path.dirname(cache_path), exist_ok=True)
            with open(cache_path, "w", encoding="utf-8") as f:
                json.dump({"ligands": self.ligands, "targets": self.targets}, f, indent=4)
        except Exception as e:
            print(f"[RegistryLoader] Failed to write cache: {e}")

    def generate_legacy_csvs(self):
        # Generate pathogen_target_registry.csv dynamically
        os.makedirs(os.path.dirname(LEGACY_TARGETS_CSV), exist_ok=True)
        
        with open(LEGACY_TARGETS_CSV, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow([
                "scenario_id", "organism_key", "target_label", "gene_symbol", "target_aliases",
                "uniprot_accession", "ncbi_protein_accession", "protein_sequence_fasta_path",
                "preferred_structure_source", "structure_source_id", "structure_source_url",
                "evidence_level", "source_traceability_id"
            ])
            
            for tgt in self.targets.values():
                # Scenario mapping
                org = tgt["organism"]
                if "Pseudomonas" in org:
                    scen = "primary_kuth_pseudomonas"
                elif "Staphylococcus" in org:
                    scen = "secondary_kuth_staphylococcus"
                else:
                    scen = f"scenario_kuth_{self.normalize_id(org.split()[0])}"
                    
                writer.writerow([
                    scen,
                    org,
                    tgt["target_name"],
                    tgt["gene"],
                    tgt["gene"].upper(),
                    tgt["uniprot"],
                    "UnknownAccession",
                    f"data/prepared/targets/{tgt['target_id']}/sequence.fasta",
                    "RCSB_PDB",
                    tgt["pdb_id"],
                    f"https://files.rcsb.org/download/{tgt['pdb_id']}.pdb" if tgt["pdb_id"] != "structure_pending" else "structure_pending",
                    "curated_literature",
                    f"UNIPROT_ACC_{tgt['uniprot']}"
                ])

        # Generate ligand_library.csv dynamically
        with open(LEGACY_LIGANDS_CSV, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow([
                "compound_id", "compound_name", "pubchem_cid", "chembl_id", "canonical_smiles",
                "molecular_formula", "molecular_weight"
            ])
            
            for lig in self.ligands.values():
                writer.writerow([
                    lig["compound_id"],
                    lig["compound_name"],
                    lig["pubchem_cid"],
                    f"CHEMBL_LEGACY_{lig['compound_id'].upper()}",
                    lig.get("canonical_smiles", "None"),
                    "Unknown",
                    "0.0"
                ])

_loader = RegistryLoader()

def get_loader() -> RegistryLoader:
    return _loader
